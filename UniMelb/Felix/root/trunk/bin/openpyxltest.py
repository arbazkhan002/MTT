'''
Created on Apr 13, 2013

@author: FelixLiu
'''
from openpyxl.reader.excel import load_workbook

wb = load_workbook(filename = r'../resources/placeNames.xlsx')

sheet_ranges = wb.get_sheet_by_name("placeNames")

print sheet_ranges.cell('D18').value